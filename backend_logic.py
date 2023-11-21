import openpyxl as xl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import datetime


def reading_excel(file_path):
    wb = xl.load_workbook(filename=file_path)

    print(wb.sheetnames, file_path)
    ws = wb['2023 Div5 Quotation Record']

    # Scanning for locating the first cell of the diagram
    i_start = -1
    signal = False
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if str(ws.cell(row=i, column=j).value).strip().lower() == 'quo no.':
                i_start = i
                signal = True
                break
        if signal:
            break

    if i_start == -1:
        print('Wrong excel format!')
        return

    # print(i_start)
    # Scanning each j columns for finding out the targets
    result_dict = {}
    for j in range(1, ws.max_column + 1):
        if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n', ' ') == 'client':
            result_dict['j_client'] = j

        if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n', ' ') == 'product name':
            result_dict['j_productName'] = j

        if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n', ' ') == 'selling price':
            result_dict['j_sellingPrice'] = j

        if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n', ' ') == 'success rate':
            result_dict['j_successRate'] = j

        if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n', ' ') == 'estimated delivery month':
            result_dict['j_month'] = j

    # print(result_dict)

    # Reading source data
    data_matrix = []
    for i in range(i_start + 1, ws.max_row + 1):
        if not ws.cell(row=i, column=result_dict['j_productName']).value:
            break

        data_line = [str(ws.cell(row=i, column=result_dict['j_client']).value).strip(),
                     str(ws.cell(row=i, column=result_dict['j_productName']).value).strip(),
                     float(ws.cell(row=i, column=result_dict['j_sellingPrice']).value) if (
                             ws.cell(row=i, column=result_dict['j_sellingPrice']).value and str(
                         ws.cell(row=i, column=result_dict['j_sellingPrice']).value).strip() != '-') else 0,
                     str(ws.cell(row=i, column=result_dict['j_successRate']).value).strip(),
                     ws.cell(row=i, column=result_dict['j_month']).value,
                     ]

        data_matrix.append(data_line)

    # print(data_matrix)

    wb.close()
    return data_matrix


def create_diagram(ws, start_row, mode):
    key_words={}
    if mode=='accept':
        key_words['title']='Secured  Business／受注案件'
        key_words['type']='Accept'
        key_words['remarks']=''
    elif mode=='reject':
        key_words['title'] = 'Reject  Business／失注案件'
        key_words['type'] = 'Reject'
        key_words['remarks'] = ''
    elif mode=='a':
        key_words['title'] = 'Opportunities A／Aヨミ案件'
        key_words['type'] = 'A'
        key_words['remarks'] = '80% can secure the business'
    elif mode=='b':
        key_words['title'] = 'Opportunities B／Bヨミ案件'
        key_words['type'] = 'B'
        key_words['remarks'] = '60% can secure the business'
    else:
        key_words['title'] = 'Opportunities C／Cヨミ案件'
        key_words['type'] = 'C'
        key_words['remarks'] = '30% can secure the business'


    # Create the first diagram
    ws.cell(row=start_row, column=2).value = key_words['title']
    font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=start_row, column=2).font = font

    ws.cell(row=start_row, column=5).value = key_words['type']
    font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=start_row, column=5).font = font

    ws.cell(row=start_row, column=6).value = key_words['remarks']
    font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=start_row, column=6).font = font

    ws.cell(row=start_row, column=17).value = 'Unit/THB'
    font = Font(name="Calibri", size=11, bold=False)
    ws.cell(row=start_row, column=17).font = font
    ws.cell(row=start_row, column=17).alignment = Alignment(horizontal='right', vertical='center')
    #start_row=3

    start_row+=1
    for j in range(2, 18):
        if j == 2:
            ws.cell(row=start_row, column=j).font = Font(name="Calibri", size=8, bold=True)
            ws.cell(row=start_row, column=j).value = 'Success rate'
        elif j == 3 or j == 4:
            ws.cell(row=start_row, column=j).font = Font(name="Calibri", size=11, bold=True)
            if j == 3:
                ws.cell(row=start_row, column=j).value = 'Client'
            else:
                ws.cell(row=start_row, column=j).value = 'Product Name'
        else:
            ws.cell(row=start_row, column=j).font = Font(name="Calibri", size=11, bold=False)

        ws.cell(row=start_row, column=j).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=start_row, column=j).fill = PatternFill("solid", fgColor="FDE9D9")
        ws.cell(row=start_row, column=j).border = Border(left=Side(border_style='thin', color='000000'),
                                                 right=Side(border_style='thin', color='000000'),
                                                 top=Side(border_style='thin', color='000000'),
                                                 bottom=Side(border_style='double', color='000000'))

    for j in range(5, 17):
        ws.cell(row=start_row, column=j).number_format = '[$-en-US]mmm-yy;@'

    ws.cell(row=start_row, column=5).value = datetime.datetime(YEAR - 1, 11, 1, 0, 0)
    ws.cell(row=start_row, column=6).value = datetime.datetime(YEAR - 1, 12, 1, 0, 0)
    ws.cell(row=start_row, column=7).value = datetime.datetime(YEAR, 1, 1, 0, 0)
    ws.cell(row=start_row, column=8).value = datetime.datetime(YEAR, 2, 1, 0, 0)
    ws.cell(row=start_row, column=9).value = datetime.datetime(YEAR, 3, 1, 0, 0)
    ws.cell(row=start_row, column=10).value = datetime.datetime(YEAR, 4, 1, 0, 0)
    ws.cell(row=start_row, column=11).value = datetime.datetime(YEAR, 5, 1, 0, 0)
    ws.cell(row=start_row, column=12).value = datetime.datetime(YEAR, 6, 1, 0, 0)
    ws.cell(row=start_row, column=13).value = datetime.datetime(YEAR, 7, 1, 0, 0)
    ws.cell(row=start_row, column=14).value = datetime.datetime(YEAR, 8, 1, 0, 0)
    ws.cell(row=start_row, column=15).value = datetime.datetime(YEAR, 9, 1, 0, 0)
    ws.cell(row=start_row, column=16).value = datetime.datetime(YEAR, 10, 1, 0, 0)

    accepted_info = []
    for data_line in data_matrix:
        if str(data_line[3]).strip() == key_words['type']:
            accepted_info.append(data_line)
    accepted_info.append([None, None, None, None, None])
    #start_row=4

    start_row+=1
    i = start_row
    for each in accepted_info:
        ws.cell(row=i, column=2).value = each[3]
        ws.cell(row=i, column=3).value = each[0]
        ws.cell(row=i, column=4).value = each[1]
        if each[4]:
            print(each[4].year, each[4].month)
            print(type(each[4].year), type(each[4].month))
            for j in range(5, 17):
                pro_year = each[4].year
                pro_month = each[4].month
                cal_year = ws.cell(row=start_row - 1, column=j).value.year
                cal_month = ws.cell(row=start_row - 1, column=j).value.month
                if pro_year == cal_year and pro_month == cal_month:
                    ws.cell(row=i, column=j).value = float(each[2])

        for j in range(2, 18):
            if 5 <= j < 17:
                ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor="FFFF00")
                ws.cell(row=i, column=j).number_format = '0.00'

            ws.cell(row=i, column=j).font = Font(name="Calibri", size=11, bold=False)
            ws.cell(row=i, column=j).border = Border(left=Side(border_style='thin', color='000000'),
                                                     right=Side(border_style='thin', color='000000'),
                                                     bottom=Side(border_style='thin', color='000000'))

        i += 1

    behind_row = i
    ws.cell(row=behind_row, column=4).value = 'Total'
    ws.cell(row=behind_row,
            column=5).value = f'=SUBTOTAL(9,E{start_row}:E{behind_row - 1})'
    ws.cell(row=behind_row,
            column=6).value = f'=SUBTOTAL(9,F{start_row}:F{behind_row - 1})'
    ws.cell(row=behind_row,
            column=7).value = f'=SUBTOTAL(9,G{start_row}:G{behind_row - 1})'
    ws.cell(row=behind_row,
            column=8).value = f'=SUBTOTAL(9,H{start_row}:H{behind_row - 1})'
    ws.cell(row=behind_row,
            column=9).value = f'=SUBTOTAL(9,I{start_row}:I{behind_row - 1})'
    ws.cell(row=behind_row,
            column=10).value = f'=SUBTOTAL(9,J{start_row}:J{behind_row - 1})'
    ws.cell(row=behind_row,
            column=11).value = f'=SUBTOTAL(9,K{start_row}:K{behind_row - 1})'
    ws.cell(row=behind_row,
            column=12).value = f'=SUBTOTAL(9,L{start_row}:L{behind_row - 1})'
    ws.cell(row=behind_row,
            column=13).value = f'=SUBTOTAL(9,M{start_row}:M{behind_row - 1})'
    ws.cell(row=behind_row,
            column=14).value = f'=SUBTOTAL(9,N{start_row}:N{behind_row - 1})'
    ws.cell(row=behind_row,
            column=15).value = f'=SUBTOTAL(9,O{start_row}:O{behind_row - 1})'
    ws.cell(row=behind_row,
            column=16).value = f'=SUBTOTAL(9,P{start_row}:P{behind_row - 1})'
    ws.cell(row=behind_row, column=17).value = f'=SUM(E{behind_row}:P{behind_row})'

    behind_row += 1
    ws.cell(row=behind_row, column=4).value = 'Quarter Total'
    ws.cell(row=behind_row, column=7).value = f'=E{behind_row - 1}+F{behind_row - 1}+G{behind_row - 1}'
    ws.cell(row=behind_row, column=10).value = f'=H{behind_row - 1}+I{behind_row - 1}+J{behind_row - 1}'
    ws.cell(row=behind_row, column=13).value = f'=K{behind_row - 1}+L{behind_row - 1}+M{behind_row - 1}'
    ws.cell(row=behind_row, column=16).value = f'=N{behind_row - 1}+O{behind_row - 1}+P{behind_row - 1}'


    behind_row += 1
    if mode=='reject' or mode=='c':
        ws.cell(row=behind_row, column=4).value = ''
        ws.cell(row=behind_row, column=7).value = ''
        ws.cell(row=behind_row, column=10).value = ''
        ws.cell(row=behind_row, column=13).value = ''
        ws.cell(row=behind_row, column=16).value = ''

    else:
        ws.cell(row=behind_row, column=4).value = 'Quarter  Achievement ratio'
        ws.cell(row=behind_row, column=7).value = f'=G{behind_row - 1}/$E$16'
        ws.cell(row=behind_row, column=10).value = f'=J{behind_row - 1}/$H$16'
        ws.cell(row=behind_row, column=13).value = f'=M{behind_row - 1}/$K$16'
        ws.cell(row=behind_row, column=16).value = f'=P{behind_row - 1}/$N$16'

    for i in range(behind_row - 2, behind_row + 1):
        for j in range(4, 18):
            ws.cell(row=i, column=j).font = Font(name="Calibri", size=11, bold=True)

    for i in range(behind_row - 2, behind_row):
        for j in range(5, 18):
            ws.cell(row=i, column=j).number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'

    for i in range(behind_row, behind_row + 1):
        for j in range(5, 18):
            ws.cell(row=i, column=j).number_format = '0%'

    #Reset start_row
    start_row=behind_row+2
    return start_row

def create_excel(data_matrix):
    wb = xl.Workbook()
    ws = wb.active

    # Setting column width
    ws.column_dimensions['A'].width = 1.89 + 0.78
    ws.column_dimensions['B'].width = 10.33 + 0.78
    ws.column_dimensions['C'].width = 11.78 + 0.78
    ws.column_dimensions['D'].width = 44.33 + 0.78
    ws.column_dimensions['E'].width = 9.22 + 0.78
    ws.column_dimensions['F'].width = 9.78 + 0.78
    ws.column_dimensions['G'].width = 9.78 + 0.78
    ws.column_dimensions['H'].width = 9.78 + 0.78
    ws.column_dimensions['I'].width = 9.78 + 0.78
    ws.column_dimensions['J'].width = 9.78 + 0.78
    ws.column_dimensions['K'].width = 9.78 + 0.78
    ws.column_dimensions['L'].width = 9.78 + 0.78
    ws.column_dimensions['M'].width = 9.78 + 0.78
    ws.column_dimensions['N'].width = 9.78 + 0.78
    ws.column_dimensions['O'].width = 9.78 + 0.78
    ws.column_dimensions['P'].width = 9.78 + 0.78
    ws.column_dimensions['Q'].width = 15.11 + 0.78

    # Create the first diagram
    ws.cell(row=1, column=2).value = '2：Opportunities Detail 案件一覧'
    font = Font(name="Calibri", size=12, bold=True)
    ws.cell(row=1, column=2).font = font

    start_row=create_diagram(ws=ws, start_row=3, mode='accept')

    start_row=create_diagram(ws=ws, start_row=start_row, mode='reject')

    start_row=create_diagram(ws=ws, start_row=start_row, mode='a')

    start_row=create_diagram(ws=ws, start_row=start_row, mode='b')

    create_diagram(ws=ws, start_row=start_row, mode='c')

    wb.save('111111111.xlsx')
    wb.close()


if __name__ == '__main__':
    YEAR = 2023
    file_path = '20230320_Quo & OS & Invo Project of 20230328test.xlsx'
    data_matrix = reading_excel(file_path=file_path)
    create_excel(data_matrix=data_matrix)
