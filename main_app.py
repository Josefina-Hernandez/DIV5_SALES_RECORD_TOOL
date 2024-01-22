# -*- coding: utf-8 -*-

from PyQt5.QtWidgets import QApplication, QWidget, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QThread, pyqtSignal

import openpyxl as xl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import sys
import os
import datetime

from ui.main import Ui_MainWindow

class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)

        self.year=datetime.datetime.now().year
        self.file_path = ''
        self.pushButton_3.clicked.connect(self.import_excel)
        self.pushButton.clicked.connect(self.export_excel)
        self.pushButton_2.clicked.connect(self.quit)
        self.pushButton_4.clicked.connect(self.about)
        self.label.setText('Excelファイルをインポートしてください…')
        self.textEdit.append('Excelファイルをインポートしてください…')
        self.textEdit.setReadOnly(True)
        self.progressBar.setMaximum(100)
        self.progressBar.setValue(0)

        current_year=datetime.datetime.now().year
        for year in range(current_year-10, current_year+11):
            self.comboBox.addItem(str(year))
        self.comboBox.setCurrentText(str(current_year))

    def import_excel(self):
        a = QFileDialog.getOpenFileName(self,
                                        'プロジェクト管理ファイルを選択してください',
                                        '',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == "":
            return
        wb = xl.load_workbook(a[0])

        judge=False
        for each in wb.sheetnames:
            if 'quotation record' in each.lower():
                judge=True
                break
        if not judge:
            self.label.setText('対象のデータシート "****** Quotation Record" が見つかりません!')
            self.textEdit.append('対象のデータシート "****** Quotation Record" が見つかりません!')
            QMessageBox.critical(self, 'フォーマットエラー', '対象のデータシート "****** Quotation Record" が見つかりません!')
            return

        wb.close()
        self.file_path=a[0]

        a_spit=a[0].split('/')
        name=a_spit[len(a_spit)-1]
        self.label.setText(f'ファイル「{name}」がインポートされました…')
        self.textEdit.append(f'ファイル「{name}」がインポートされました…')
        self.update_progressbar(0, 100)

    def export_excel(self):
        if self.file_path=='':
            self.label.setText('まずExcelファイルをインポートしてください！')
            self.textEdit.append('まずExcelファイルをインポートしてください！')
            QMessageBox.critical(self, 'Excelファイルなし',
                                 'まずExcelファイルをインポートしてください！')
            return

        a = QFileDialog.getSaveFileName(self,
                                        'Please select the OT sheet file path.',
                                        f'./Opportunities Detail 案件一覧{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}',
                                        'Excel Files (*.xlsx);;All Files (*)')

        if a[0] == '':
            self.label.setText('Excel出力をキャンセルしました。')
            self.textEdit.append('Excel出力をキャンセルしました。')
            return

        wb = xl.Workbook()
        try:
            wb.save(filename=a[0])
        except PermissionError:
            QMessageBox.critical(self, '許可が拒否されました！',
                                 '許可が拒否されました！ Please close the excel file with the same filename first!')
            wb.close()
            return

        wb.close()

        self.working = Working(import_filename=self.file_path, export_filename=a[0], YEAR=int(self.comboBox.currentText()))
        self.working.finish_msgbox.connect(self.finish_msgbox)
        self.working.update_msg.connect(self.update_msg)
        self.working.update_progressbar.connect(self.update_progressbar)


        self.working.start()

    def finish_msgbox(self, title, text):
        QMessageBox.information(self, title, text)

    def update_msg(self, msg):
        self.label.setText(msg)
        self.textEdit.append(msg)

    def update_progressbar(self, percentage, max):
        self.progressBar.setMaximum(max)
        self.progressBar.setValue(percentage)

    def quit(self):
        a = QMessageBox.question(self, '操作確認', 'システムを閉じますか？', QMessageBox.Yes | QMessageBox.No)
        if a != 16384:
            return

        # print(12345)
        sys.exit()

    def closeEvent(self, event):
        a = QMessageBox.question(self, '操作確認', 'システムを閉じますか？', QMessageBox.Yes | QMessageBox.No)
        if a != 16384:
            event.ignore()
            return

        sys.exit()

    def about(self):
        QMessageBox.information(self, 'ツール情報',
                                'AKT DIV.5 見積管理ツール (V1.5版)\n'
                                '言語: 日本語\n'
                                'バージョン: V1.5版\n\n'
                                '開発者：An Lu\n'
                                '開発時間：2023年3月29日\n\n'
                                '連絡先: (+66)84-208-1862\n'
                                'E-mail: lu@akaganethailand.co.th\n'
                                '住所：16 Compomax Building, 5th Floor, Room No. 502, Soi Ekamai 4, Sukhumvit 63 Rd., Prakanongnua, Vadhana, Bangkok 10110 (Head office)\n\n'
                                'AKAGANE(THAILAND) CO., LTD.')

class Working(QThread):
    finish_msgbox = pyqtSignal(str, str)
    update_msg = pyqtSignal(str)
    update_progressbar = pyqtSignal(int, int)


    def __init__(self,import_filename, export_filename, YEAR):
        super(Working, self).__init__()
        self.import_filename=import_filename
        self.export_filename=export_filename
        self.YEAR = YEAR

    def reading_excel(self, file_path):
        wb = xl.load_workbook(filename=file_path, data_only=True)  #updated on 7/20/2023 for v1.3, added "data_only=True"
        #ws = wb[f'{self.YEAR} Div5 Quotation Record']

        for each in wb.sheetnames:
            if 'quotation record' in each.lower():
                ws = wb[f'{each}']
                break

        # Scanning for locating the first cell of the diagram
        self.update_msg.emit('')
        i_start = -1
        signal = False
        for i in range(1, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                if str(ws.cell(row=i, column=j).value).strip().lower() == 'quo no.':
                    i_start = i
                    signal = True
                    break
            if signal:
                break

        if i_start == -1:
            #print('Wrong excel format!')
            return -1

        # print(i_start)
        # Scanning each j columns for finding out the targets

        self.update_msg.emit('インポートファイルの目標シートと題名を検索しています…')
        self.update_progressbar.emit(50, 100)

        result_dict = {}
        for j in range(1, ws.max_column + 1):
            if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n', ' ') == 'client':
                result_dict['j_client'] = j

            if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n', ' ') == 'product name':
                result_dict['j_productName'] = j

            if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n', ' ') == 'selling price':
                result_dict['j_sellingPrice'] = j

            if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n', ' ') == 'success rate':
                result_dict['j_successRate'] = j

            if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n',
                                                                                 ' ') == 'estimated delivery month':
                result_dict['j_month'] = j

            if str(ws.cell(row=i_start, column=j).value).strip().lower().replace('\n',
                                                                                 ' ') == 'status':
                result_dict['j_status'] = j



        # print(result_dict)
        self.update_msg.emit(f'目標題名を発見しました：{result_dict}…')

        # Reading source data
        data_matrix = []
        for i in range(i_start + 1, ws.max_row + 1):
            if not ws.cell(row=i, column=result_dict['j_productName']).value:
                break

            data_line = [str(ws.cell(row=i, column=result_dict['j_client']).value).strip(),
                         str(ws.cell(row=i, column=result_dict['j_productName']).value).strip(),
                         float(ws.cell(row=i, column=result_dict['j_sellingPrice']).value) if (
                                 ws.cell(row=i, column=result_dict['j_sellingPrice']).value and str(
                             ws.cell(row=i, column=result_dict['j_sellingPrice']).value).strip() != '-') else 0,
                         str(ws.cell(row=i, column=result_dict['j_successRate']).value).strip(),
                         ws.cell(row=i, column=result_dict['j_month']).value,
                         str(ws.cell(row=i, column=result_dict['j_status']).value).strip().lower(),
                         ]

            self.update_msg.emit(f'ソースデータを読み込んでいます：{data_line}…')
            self.update_progressbar.emit(i, ws.max_row)
            data_matrix.append(data_line)

        # print(data_matrix)

        wb.close()
        return data_matrix

    def create_diagram(self, ws, start_row, mode, data_matrix):
        key_words = {}
        if mode == 'accept':
            key_words['title'] = 'Secured  Business／受注案件'
            key_words['type'] = 'Accept'
            key_words['remarks'] = ''
        elif mode == 'reject':
            key_words['title'] = 'Reject  Business／失注案件'
            key_words['type'] = 'Reject'
            key_words['remarks'] = ''
        elif mode == 'a':
            key_words['title'] = 'Opportunities A／Aヨミ案件'
            key_words['type'] = 'A'
            key_words['remarks'] = '80% can secure the business'
        elif mode == 'b':
            key_words['title'] = 'Opportunities B／Bヨミ案件'
            key_words['type'] = 'B'
            key_words['remarks'] = '60% can secure the business'
        else:
            key_words['title'] = 'Opportunities C／Cヨミ案件'
            key_words['type'] = 'C'
            key_words['remarks'] = '30% can secure the business'

        # Create the first diagram
        ws.cell(row=start_row, column=2).value = key_words['title']
        font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=start_row, column=2).font = font

        ws.cell(row=start_row, column=5).value = key_words['type']
        font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=start_row, column=5).font = font

        ws.cell(row=start_row, column=6).value = key_words['remarks']
        font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=start_row, column=6).font = font

        ws.cell(row=start_row, column=17).value = 'Unit/THB'
        font = Font(name="Calibri", size=11, bold=False)
        ws.cell(row=start_row, column=17).font = font
        ws.cell(row=start_row, column=17).alignment = Alignment(horizontal='right', vertical='center')
        # start_row=3

        start_row += 1
        for j in range(2, 18):
            if j == 2:
                ws.cell(row=start_row, column=j).font = Font(name="Calibri", size=8, bold=True)
                ws.cell(row=start_row, column=j).value = 'Success rate'
            elif j == 3 or j == 4:
                ws.cell(row=start_row, column=j).font = Font(name="Calibri", size=11, bold=True)
                if j == 3:
                    ws.cell(row=start_row, column=j).value = 'Client'
                else:
                    ws.cell(row=start_row, column=j).value = 'Product Name'
            else:
                ws.cell(row=start_row, column=j).font = Font(name="Calibri", size=11, bold=False)

            ws.cell(row=start_row, column=j).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=start_row, column=j).fill = PatternFill("solid", fgColor="FDE9D9")
            ws.cell(row=start_row, column=j).border = Border(left=Side(border_style='thin', color='000000'),
                                                             right=Side(border_style='thin', color='000000'),
                                                             top=Side(border_style='thin', color='000000'),
                                                             bottom=Side(border_style='double', color='000000'))

        for j in range(5, 17):
            ws.cell(row=start_row, column=j).number_format = '[$-en-US]mmm-yy;@'

        ws.cell(row=start_row, column=5).value = datetime.datetime(self.YEAR - 1, 11, 1, 0, 0)
        ws.cell(row=start_row, column=6).value = datetime.datetime(self.YEAR - 1, 12, 1, 0, 0)
        ws.cell(row=start_row, column=7).value = datetime.datetime(self.YEAR, 1, 1, 0, 0)
        ws.cell(row=start_row, column=8).value = datetime.datetime(self.YEAR, 2, 1, 0, 0)
        ws.cell(row=start_row, column=9).value = datetime.datetime(self.YEAR, 3, 1, 0, 0)
        ws.cell(row=start_row, column=10).value = datetime.datetime(self.YEAR, 4, 1, 0, 0)
        ws.cell(row=start_row, column=11).value = datetime.datetime(self.YEAR, 5, 1, 0, 0)
        ws.cell(row=start_row, column=12).value = datetime.datetime(self.YEAR, 6, 1, 0, 0)
        ws.cell(row=start_row, column=13).value = datetime.datetime(self.YEAR, 7, 1, 0, 0)
        ws.cell(row=start_row, column=14).value = datetime.datetime(self.YEAR, 8, 1, 0, 0)
        ws.cell(row=start_row, column=15).value = datetime.datetime(self.YEAR, 9, 1, 0, 0)
        ws.cell(row=start_row, column=16).value = datetime.datetime(self.YEAR, 10, 1, 0, 0)

        accepted_info = []
        for data_line in data_matrix:
            if data_line[5].strip()=='reject':
                data_line[3]='Reject'
            elif data_line[5].strip()=='accept':
                data_line[3]='Accept'

            if str(data_line[3]).strip() == key_words['type']:
                accepted_info.append(data_line)
        accepted_info.append([None, None, None, None, None])
        # start_row=4

        start_row += 1
        i = start_row
        for each in accepted_info:
            ws.cell(row=i, column=2).value = each[3]
            ws.cell(row=i, column=3).value = each[0]
            ws.cell(row=i, column=4).value = each[1]
            if each[4] and isinstance(each[4], datetime.datetime):
                print(isinstance(each[4], datetime.datetime))
                print(each[4].year, each[4].month)
                print(type(each[4].year), type(each[4].month))
                for j in range(5, 17):
                    pro_year = each[4].year
                    pro_month = each[4].month
                    cal_year = ws.cell(row=start_row - 1, column=j).value.year
                    cal_month = ws.cell(row=start_row - 1, column=j).value.month
                    if pro_year == cal_year and pro_month == cal_month:
                        ws.cell(row=i, column=j).value = float(each[2])

            for j in range(2, 18):
                if 5 <= j < 17:
                    ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor="FFFF00")
                    ws.cell(row=i, column=j).number_format = '###,###'

                ws.cell(row=i, column=j).font = Font(name="Calibri", size=11, bold=False)
                ws.cell(row=i, column=j).border = Border(left=Side(border_style='thin', color='000000'),
                                                         right=Side(border_style='thin', color='000000'),
                                                         bottom=Side(border_style='thin', color='000000'))

            i += 1

        behind_row = i
        ws.cell(row=behind_row, column=4).value = 'AKJ Total'
        ws.cell(row=behind_row,
                column=5).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",E{start_row}:E{behind_row - 1})'
        ws.cell(row=behind_row,
                column=6).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",F{start_row}:F{behind_row - 1})'
        ws.cell(row=behind_row,
                column=7).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",G{start_row}:G{behind_row - 1})'
        ws.cell(row=behind_row,
                column=8).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",H{start_row}:H{behind_row - 1})'
        ws.cell(row=behind_row,
                column=9).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",I{start_row}:I{behind_row - 1})'
        ws.cell(row=behind_row,
                column=10).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",J{start_row}:J{behind_row - 1})'
        ws.cell(row=behind_row,
                column=11).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",K{start_row}:K{behind_row - 1})'
        ws.cell(row=behind_row,
                column=12).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",L{start_row}:L{behind_row - 1})'
        ws.cell(row=behind_row,
                column=13).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",M{start_row}:M{behind_row - 1})'
        ws.cell(row=behind_row,
                column=14).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",N{start_row}:N{behind_row - 1})'
        ws.cell(row=behind_row,
                column=15).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",O{start_row}:O{behind_row - 1})'
        ws.cell(row=behind_row,
                column=16).value = f'=SUMIF($C${start_row}:$P${behind_row - 1},"AKJ",P{start_row}:P{behind_row - 1})'
        ws.cell(row=behind_row, column=17).value = f'=SUM(E{behind_row}:P{behind_row})'

        behind_row += 1
        ws.cell(row=behind_row, column=4).value = 'CB Total'
        ws.cell(row=behind_row, column=5).value = f'=SUM(E{start_row}:E{behind_row - 2})-E{behind_row - 1}'
        ws.cell(row=behind_row, column=6).value = f'=SUM(F{start_row}:F{behind_row - 2})-F{behind_row - 1}'
        ws.cell(row=behind_row, column=7).value = f'=SUM(G{start_row}:G{behind_row - 2})-G{behind_row - 1}'
        ws.cell(row=behind_row, column=8).value = f'=SUM(H{start_row}:H{behind_row - 2})-H{behind_row - 1}'
        ws.cell(row=behind_row, column=9).value = f'=SUM(I{start_row}:I{behind_row - 2})-I{behind_row - 1}'
        ws.cell(row=behind_row, column=10).value = f'=SUM(J{start_row}:J{behind_row - 2})-J{behind_row - 1}'
        ws.cell(row=behind_row, column=11).value = f'=SUM(K{start_row}:K{behind_row - 2})-K{behind_row - 1}'
        ws.cell(row=behind_row, column=12).value = f'=SUM(L{start_row}:L{behind_row - 2})-L{behind_row - 1}'
        ws.cell(row=behind_row, column=13).value = f'=SUM(M{start_row}:M{behind_row - 2})-M{behind_row - 1}'
        ws.cell(row=behind_row, column=14).value = f'=SUM(N{start_row}:N{behind_row - 2})-N{behind_row - 1}'
        ws.cell(row=behind_row, column=15).value = f'=SUM(O{start_row}:O{behind_row - 2})-O{behind_row - 1}'
        ws.cell(row=behind_row, column=16).value = f'=SUM(P{start_row}:P{behind_row - 2})-P{behind_row - 1}'
        ws.cell(row=behind_row, column=17).value = f'=SUBTOTAL(9,E{behind_row}:P{behind_row})'

        behind_row += 1
        ws.cell(row=behind_row, column=4).value = 'Quarter Total'
        ws.cell(row=behind_row, column=7).value = f'=SUM(E{behind_row - 2}:G{behind_row - 1})'
        ws.cell(row=behind_row, column=10).value = f'=SUM(H{behind_row - 2}:J{behind_row - 1})'
        ws.cell(row=behind_row, column=13).value = f'=SUM(K{behind_row - 2}:M{behind_row - 1})'
        ws.cell(row=behind_row, column=16).value = f'=SUM(N{behind_row - 2}:P{behind_row - 1})'

        behind_row += 1

        global REFER_ROW
        if mode == 'reject' or mode == 'c':
            ws.cell(row=behind_row, column=4).value = ''
            ws.cell(row=behind_row, column=7).value = ''
            ws.cell(row=behind_row, column=10).value = ''
            ws.cell(row=behind_row, column=13).value = ''
            ws.cell(row=behind_row, column=16).value = ''

        elif mode == 'a':
            ws.cell(row=behind_row, column=4).value = 'Quarter  Achievement ratio'
            ws.cell(row=behind_row, column=7).value = f'=G{behind_row - 1}/(E17+E10)'
            ws.cell(row=behind_row, column=10).value = f'=J{behind_row - 1}/(H17+H10)'
            ws.cell(row=behind_row, column=13).value = f'=M{behind_row - 1}/(K17+K10)'
            ws.cell(row=behind_row, column=16).value = f'=P{behind_row - 1}/(N17+N10)'

            ws.cell(row=behind_row, column=17).value = f'=(Q{behind_row - 3}*0.8+Q6+Q13)/(Q10+Q17)'

            REFER_ROW = behind_row - 1

        elif mode == 'b':
            ws.cell(row=behind_row, column=4).value = 'Quarter  Achievement ratio'
            ws.cell(row=behind_row, column=7).value = f'=(G{behind_row - 1}*0.6+G{REFER_ROW}*0.8)/(E17+E10)'
            ws.cell(row=behind_row, column=10).value = f'=(J{behind_row - 1}*0.6+J{REFER_ROW}*0.8)/(H17+H10)'
            ws.cell(row=behind_row, column=13).value = f'=(M{behind_row - 1}*0.6+M{REFER_ROW}*0.8)/(K17+K10)'
            ws.cell(row=behind_row, column=16).value = f'=(P{behind_row - 1}*0.6+P{REFER_ROW}*0.8)/(N17+N10)'

            ws.cell(row=behind_row, column=17).value = f'=(Q{REFER_ROW}*0.8+Q{behind_row - 1}*0.6+Q6+Q13)/(Q10+Q17)'

        else:
            ws.cell(row=behind_row, column=4).value = 'Quarter  Achievement ratio'
            ws.cell(row=behind_row, column=7).value = f'=G{behind_row - 1}/(E17+E10)'
            ws.cell(row=behind_row, column=10).value = f'=J{behind_row - 1}/(H17+H10)'
            ws.cell(row=behind_row, column=13).value = f'=M{behind_row - 1}/(K17+K10)'
            ws.cell(row=behind_row, column=16).value = f'=P{behind_row - 1}/(N17+N10)'

        for i in range(behind_row - 3, behind_row + 1):
            for j in range(4, 18):
                ws.cell(row=i, column=j).font = Font(name="Calibri", size=11, bold=True)

        for i in range(behind_row - 3, behind_row):
            for j in range(5, 18):
                ws.cell(row=i, column=j).number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'

        for i in range(behind_row, behind_row + 1):
            for j in range(5, 18):
                ws.cell(row=i, column=j).number_format = '0%'

        # Reset start_row
        start_row = behind_row + 2
        return start_row

    def create_excel(self, data_matrix):

        self.update_msg.emit('Excelファイルを出力しています…')
        self.update_progressbar.emit(0, 100)
        wb = xl.Workbook()
        ws = wb.active

        self.update_msg.emit('Excel書式を調整しています…')
        self.update_progressbar.emit(10, 100)
        # Setting column width
        ws.column_dimensions['A'].width = 1.89 + 0.78
        ws.column_dimensions['B'].width = 10.33 + 0.78
        ws.column_dimensions['C'].width = 11.78 + 0.78
        ws.column_dimensions['D'].width = 44.33 + 0.78
        ws.column_dimensions['E'].width = 9.22 + 0.78
        ws.column_dimensions['F'].width = 9.78 + 0.78
        ws.column_dimensions['G'].width = 9.78 + 0.78
        ws.column_dimensions['H'].width = 9.78 + 0.78
        ws.column_dimensions['I'].width = 9.78 + 0.78
        ws.column_dimensions['J'].width = 9.78 + 0.78
        ws.column_dimensions['K'].width = 9.78 + 0.78
        ws.column_dimensions['L'].width = 9.78 + 0.78
        ws.column_dimensions['M'].width = 9.78 + 0.78
        ws.column_dimensions['N'].width = 9.78 + 0.78
        ws.column_dimensions['O'].width = 9.78 + 0.78
        ws.column_dimensions['P'].width = 9.78 + 0.78
        ws.column_dimensions['Q'].width = 15.11 + 0.78

        # Create the first diagram
        ws.cell(row=1, column=2).value = '2：Opportunities Detail 案件一覧'
        font = Font(name="Calibri", size=12, bold=True)
        ws.cell(row=1, column=2).font = font

        self.update_msg.emit('受注案件表を作成しています…')
        self.update_progressbar.emit(20, 100)
        start_row = self.create_diagram(ws=ws, start_row=3, mode='accept', data_matrix=data_matrix)

        self.update_msg.emit('失注案件表を作成しています…')
        self.update_progressbar.emit(30, 100)
        start_row = self.create_diagram(ws=ws, start_row=start_row, mode='reject', data_matrix=data_matrix)

        self.update_msg.emit('Aタイプ案件表を作成しています…')
        self.update_progressbar.emit(40, 100)
        start_row = self.create_diagram(ws=ws, start_row=start_row, mode='a', data_matrix=data_matrix)

        self.update_msg.emit('Bタイプ案件表を作成しています…')
        self.update_progressbar.emit(70, 100)
        start_row = self.create_diagram(ws=ws, start_row=start_row, mode='b', data_matrix=data_matrix)

        self.update_msg.emit('Cタイプ案件表を作成しています…')
        self.update_progressbar.emit(80, 100)
        self.create_diagram(ws=ws, start_row=start_row, mode='c', data_matrix=data_matrix)

        self.update_msg.emit(F'Excelファイル「{self.export_filename}」を保存しています…')
        self.update_progressbar.emit(100, 100)
        wb.save(self.export_filename)
        wb.close()

    def run(self):
        self.update_progressbar.emit(0, 100)
        file_path = self.import_filename
        self.update_msg.emit('Excelファイルを読み込んでいます…')
        data_matrix = self.reading_excel(file_path=file_path)
        if data_matrix==-1:
            self.update_msg.emit('インポートされたファイル形式は違います！正しいExcelファイルをインポートしてください…')
            self.finish_msgbox.emit('エラー', 'インポートされたファイル形式は違います！正しいExcelファイルをインポートしてください…')
            self.update_progressbar.emit(100, 100)
            os.remove(self.export_filename)
            return

        self.create_excel(data_matrix=data_matrix)

        self.update_msg.emit(F'案件表の作成が完了しました…')
        self.update_progressbar.emit(100, 100)
        self.finish_msgbox.emit('完了', '案件表の作成が完了しました…')
        MainWindow.file_path=''


if __name__ == '__main__':
    app = QApplication(sys.argv)

    REFER_ROW=-1

    MainWindow = MainWindow()
    MainWindow.show()
    sys.exit(app.exec_())